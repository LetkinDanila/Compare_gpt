import json
import pandas as pd
import os
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
from requests.exceptions import RequestException
from dataclasses import dataclass, field
from typing import List

HEADER_TOKEN = "Bearer 465d457953dead9047bc2343d4c6d7a897fdd467f9a74f9b"

def select_file():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        filetypes=[("JSON files", "*.json")],
        title="Select a JSON file"
    )
    return file_path

def load_data(filename):
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Error loading JSON file: {e}")
        return {}

def modify_prompt(promt, params):
    for key, value in params.items():
        promt = promt.replace(f"{{{key}}}", value)
    return promt

def parse_models(data):
    models = data['models']
    promt_template = data['prompt']  
    tests = data['tests'] 
    
    result = []
    
    for model in models:
        for test in tests:
            modified_promt = modify_prompt(promt_template, test['params'])
            
            model_object = {
                "prompt": modified_promt,
                "model": model,
                "input": test.get('input', ''),
                "expected": test.get('expected', ''),
                "params": test.get('params', {}),
                "max_tokens": test.get('max_tokens', 100),
                "temperature": test.get('temperature', 0),
                "status": "Not completed",
                "response": " ",
                "response_time": 0
            }
            result.append(model_object)
    
    return result

def compare_results(data):
    for item in data:
        if item['expected'] == item['response']:
            item['status'] = "passed"
        else:
            item['status'] = "failed"
    return data


def write_to_excel(data, file_path):
    file_path = file_path.rstrip(".json")
    file_path = file_path + ".xlsx"
    rows = []
    current_model = None
    
    for item in data:
        if item['model'] != current_model:
            if current_model is not None:
                rows.append({})
            current_model = item['model']
        
        rows.append({
            'model': item['model'],
            'prompt': item['prompt'],
            'input': item['input'],
            'expected': item['expected'],
            'response': item['response'],
            'response_time': item['response_time'],
            'status': item['status']
        })
    
    df = pd.DataFrame(rows)
    
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    df.to_excel(file_path, index=False)

    wb = load_workbook(file_path)
    ws = wb.active

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        status_cell = row[6]
        if status_cell.value == 'passed':
            status_cell.fill = green_fill
        elif status_cell.value == 'failed':
            status_cell.fill = red_fill
    
    wb.save(file_path)

def data2messageGPT(data):
    msg = {
        "model": data.get("model", "default-model"),
        "max_tokens": data.get("max_tokens", 100),
        "temperature": data.get("temperature", 0),
        "messages": [
            {
                "role": "system",
                "content": data.get("prompt", "")  # Значение из data или пустая строка по умолчанию
            },
            {
                "role": "user",
                "content": data.get("input", "")  # Значение из data или пустая строка по умолчанию
            }
        ]
    }
    msg = json.dumps(msg)
    return msg

def data2messageGigachat(data):
    msg = {
        #"prompt": data.get("input", ""),
        #"system": data.get("promt", ""),
        "prompt": data.get("promt", ""),
        "system": data.get("input", ""),
        "max_tokens": data.get("max_tokens", 100),
        "temperature": data.get("temperature", 0)
    }
    msg = json.dumps(msg)
    return msg

def send_gpt_request(msg):
    url = "https://paidmethods.mcn.ru/api/protected/api/llm/chat_completion_raw"
    headers = {
        "Authorization": HEADER_TOKEN,
        "Content-Type": "application/json"
    }

    result = {
        "response": " ",
        "response_time": 0
    }

    try:
        response = requests.post(url, headers=headers, data=msg)
        response_data = response.json()
        if "completion" in response_data:
            result['response'] = response_data["completion"]
            result['response_time'] = response.elapsed.total_seconds()
        else:
            result['response'] = "Error: No completion field"
    except RequestException as e:
        result['response'] = f"Error: {e}"
    
    return result

def send_gigachat_request(msg):
    url = "https://paidmethods.mcn.ru/api/protected/api/llm/gigachat/chat_completion"
    headers = {
        "Authorization": HEADER_TOKEN,
        "Content-Type": "application/json"
    }

    result = {
        "response": " ",
        "response_time": 0
    }
    try:
        response = requests.post(url, headers=headers, data=msg)
        response_data = response.json()
        if "completion" in response_data:
            result['response'] = response_data["completion"]
            result['response_time'] = response.elapsed.total_seconds()
        else:
            result['response'] = "Error: No completion field"
    except RequestException as e:
        result['response'] = f"Error: {e}"

    return result

def send_request(data):
    for item in data:
        if item['model'] in ["gpt-4o", "gpt-4o-mini"]:
            msg = data2messageGPT(item)
            result = send_gpt_request(msg)
        elif item['model'] == "gigachat":
            msg = data2messageGigachat(item)
            result = send_gigachat_request(msg)
        else:
            result = {"response": "Unknown model", "response_time": 0}

        item['response'] = result['response']
        item['response_time'] = result['response_time']

    return data

file_path = select_file()
#res_data = parse_models(load_data(file_path))
#print(res_data)
res_data = compare_results(send_request(parse_models(load_data(file_path))))
write_to_excel(res_data, file_path)